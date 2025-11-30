from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, abort
import pandas as pd
import os
import pickle
from werkzeug.utils import secure_filename
import pyodbc
import datetime
from dotenv import load_dotenv
import google.generativeai as genai
from functools import lru_cache
import numpy as np
from datetime import date
import shap
import matplotlib
import base64
from email.message import EmailMessage
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
from authlib.integrations.flask_client import OAuth
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from secrets import token_urlsafe
import random
import re
import string
import re
matplotlib.use('Agg')
import matplotlib.pyplot as plt

# ==========================================
# Cấu hình Flask
# ==========================================
load_dotenv()
app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "cvdapp-secret-key")

oauth = OAuth(app)
SOCIAL_PROVIDERS = {"google": False}
DEFAULT_SOCIAL_PASSWORD = os.getenv("DEFAULT_SOCIAL_PASSWORD", "123456")
GMAIL_PATTERN = re.compile(r"^[A-Za-z0-9._%+-]+@gmail\.com$", re.IGNORECASE)
INVITE_TOKEN_SALT = "patient_invite_token"
INVITE_TOKEN_MAX_AGE = 48 * 3600


def is_valid_gmail(email: str) -> bool:
    return bool(email and GMAIL_PATTERN.match(email.strip()))


def _invite_serializer():
    secret = app.secret_key or os.getenv("SECRET_KEY", "cvdapp-secret-key")
    return URLSafeTimedSerializer(secret, salt=INVITE_TOKEN_SALT)


def generate_invite_token(data: dict) -> str:
    return _invite_serializer().dumps(data)


def decode_invite_token(token: str, max_age: int = INVITE_TOKEN_MAX_AGE) -> dict:
    return _invite_serializer().loads(token, max_age=max_age)

GOOGLE_CLIENT_ID = os.getenv("GOOGLE_CLIENT_ID")
GOOGLE_CLIENT_SECRET = os.getenv("GOOGLE_CLIENT_SECRET")

if GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET:
    oauth.register(
        name="google",
        client_id=GOOGLE_CLIENT_ID,
        client_secret=GOOGLE_CLIENT_SECRET,
        server_metadata_url="https://accounts.google.com/.well-known/openid-configuration",
        client_kwargs={"scope": "openid email profile"},
    )
    SOCIAL_PROVIDERS["google"] = True

GMAIL_CLIENT_ID = os.getenv("GMAIL_CLIENT_ID_SEND")
GMAIL_CLIENT_SECRET = os.getenv("GMAIL_CLIENT_SECRET_SEND")
GMAIL_REFRESH_TOKEN = os.getenv("GMAIL_REFRESH_TOKEN")
GMAIL_SENDER = os.getenv("GMAIL_SENDER")
GMAIL_SCOPES = ["https://www.googleapis.com/auth/gmail.send"]

def _build_gmail_service():
    if not (GMAIL_CLIENT_ID and GMAIL_CLIENT_SECRET and GMAIL_REFRESH_TOKEN and GMAIL_SENDER):
        raise RuntimeError("Chưa cấu hình Gmail API (client id/secret, refresh token, sender).")
    creds = Credentials(
        None,
        refresh_token=GMAIL_REFRESH_TOKEN,
        token_uri="https://oauth2.googleapis.com/token",
        client_id=GMAIL_CLIENT_ID,
        client_secret=GMAIL_CLIENT_SECRET,
        scopes=GMAIL_SCOPES,
    )
    return build("gmail", "v1", credentials=creds, cache_discovery=False)


def send_email(to_email: str, subject: str, html_body: str):
    """Gửi email HTML thông qua Gmail API."""
    service = _build_gmail_service()

    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = GMAIL_SENDER
    message["To"] = to_email
    message.set_content("Trình duyệt email của bạn không hỗ trợ nội dung HTML.")
    message.add_alternative(html_body, subtype="html")

    raw = base64.urlsafe_b64encode(message.as_bytes()).decode()
    service.users().messages().send(userId="me", body={"raw": raw}).execute()

def get_connection():
    return pyodbc.connect(
        "DRIVER={ODBC Driver 18 for SQL Server};"
        "SERVER=PC1\\LNTUANDAT;"
        "DATABASE=CVD_App;"
        "Trusted_Connection=yes;"
        "Encrypt=yes;"
        "TrustServerCertificate=yes;"
    )
# def get_connection():
#     return pyodbc.connect(
#         "DRIVER={SQL Server};"
#         "SERVER=HKT;"
#         "DATABASE=CVD_App;"
#         "UID=sa;"
#         "PWD=123"
#     )
@app.context_processor
def inject_social_flags():
    return {
        "social_google_enabled": SOCIAL_PROVIDERS.get("google", False),
    }

# ==========================================
# Cấu hình Gemini AI
# ==========================================
MODEL_NAME = "models/gemini-2.5-flash-lite"
genai.configure(api_key=os.getenv("GEMINI_API_KEY"))

@lru_cache(maxsize=128)
def get_ai_advice_cached(prompt: str) -> str:
    try:
        model = genai.GenerativeModel(MODEL_NAME)
        response = model.generate_content(prompt)
        return response.text
    except Exception as e:
        return f"Không thể lấy lời khuyên AI: {e}"

# ==========================================
# Load mô hình XGBoost
# ==========================================
xgb_model = None
XGB_FEATURE_ORDER = [
    "age", "gender", "height", "weight",
    "ap_hi", "ap_lo", "cholesterol", "gluc",
    "smoke", "alco", "active",
]
try:
    import xgboost as xgb
    MODEL_PATH = "xgb_final_model.pkl"
    if os.path.exists(MODEL_PATH):
        with open(MODEL_PATH, "rb") as f:
            xgb_model = pickle.load(f)
        fresh_params = xgb.XGBClassifier().get_params()
        for attr, value in fresh_params.items():
            if not hasattr(xgb_model, attr):
                setattr(xgb_model, attr, value)
        setattr(xgb_model, "use_label_encoder", False)
        print("Mô hình XGBoost (.pkl) đã load thành công.")
    else:
        print("Không tìm thấy file mô hình (.pkl), sẽ dùng heuristic.")
except Exception as e:
    print(f"Không thể load mô hình XGBoost: {e}")
    xgb_model = None

# Warm up mô hình ngay khi Flask khởi động
@app.before_request
def warmup_model():
    """Chạy warm-up 1 lần duy nhất khi nhận request đầu tiên."""
    if not getattr(app, "_model_warmed", False):
        try:
            import numpy as np, shap
            dummy = np.array([[50, 1, 170, 70, 120, 80, 2, 1, 0, 0, 1]])
            _ = xgb_model.predict_proba(dummy)
            shap.TreeExplainer(xgb_model)
            print("Warm-up hoàn tất, model & SHAP đã cache.")
            app._model_warmed = True  # đánh dấu đã warm-up rồi
        except Exception as e:
            print(f"Warm-up model lỗi: {e}")

# ==========================================
# Cấu hình upload
# ==========================================
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

import re

def _clear_pending_registration():
    session.pop('pending_registration', None)
    session.pop('pending_registration_code', None)
    session.pop('pending_registration_exp', None)

def _send_verification_email(name, email, code):
    html_body = render_template(
        'emails/verification_code.html',
        name=name or 'bạn',
        code=code
    )
    send_email(
        email,
        "Mã xác nhận đăng ký CVD-App",
        html_body
    )

# ==========================================
# 🧬 Đăng ký tài khoản
# ==========================================
@app.route('/register', methods=['GET', 'POST'])
def register():
    today = date.today().strftime('%Y-%m-%d')

    if request.method == 'POST':
        ho_ten = request.form.get('ho_ten')
        gioi_tinh = request.form.get('gioi_tinh')
        ngay_sinh = request.form.get('ngay_sinh')
        email = request.form.get('email').strip().lower()
        mat_khau = request.form.get('mat_khau')
        mat_khau_confirm = request.form.get('mat_khau_confirm')

        if mat_khau != mat_khau_confirm:
            flash('Mật khẩu không khớp.', 'warning')
            return render_template('register.html', today=today)

        role = 'patient'

        if ngay_sinh:
            try:
                birth_date = datetime.datetime.strptime(ngay_sinh, "%Y-%m-%d").date()
                age = (date.today() - birth_date).days // 365
                if age < 16:
                    flash("Tuổi phải từ 16 trở lên.", "warning")
                    return render_template('register.html', today=today)
            except ValueError:
                flash("Ngày sinh không hợp lệ.", "warning")
                return render_template('register.html', today=today)
        else:
            flash("Vui lòng nhập ngày sinh.", "warning")
            return render_template('register.html', today=today)

        if not re.match(r'^(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{8,}$', mat_khau):
            flash('Mật khẩu cần ít nhất 8 ký tự, gồm chữ hoa, số và ký tự đặc biệt.', 'warning')
            return render_template('register.html', today=today)

        conn = get_connection()
        cur = conn.cursor()
        cur.execute('SELECT ID FROM NguoiDung WHERE Email = ?', (email,))
        if cur.fetchone():
            conn.close()
            flash('Email đã tồn tại! Vui lòng chọn email khác.', 'warning')
            return render_template('register.html', today=today)
        conn.close()

        verification_code = f"{random.randint(100000, 999999)}"
        session['pending_registration'] = {
            'ho_ten': ho_ten,
            'gioi_tinh': gioi_tinh,
            'ngay_sinh': ngay_sinh or None,
            'email': email,
            'mat_khau': mat_khau,
            'role': role
        }
        session['pending_registration_code'] = verification_code
        session['pending_registration_exp'] = (
            datetime.datetime.utcnow() + datetime.timedelta(minutes=10)
        ).isoformat()

        try:
            _send_verification_email(ho_ten, email, verification_code)
            flash('Đã gửi mã xác nhận đến email của bạn. Vui lòng kiểm tra và nhập mã để hoàn tất đăng ký.', 'info')
            return redirect(url_for('verify_email'))
        except Exception as e:
            _clear_pending_registration()
            flash(f'Không thể gửi email xác nhận: {e}', 'danger')
            return render_template('register.html', today=today)

    return render_template('register.html', today=today)

# ==========================================
# Xác thực email đăng ký
# ==========================================
@app.route('/verify-email', methods=['GET', 'POST'])
def verify_email():
    pending = session.get('pending_registration')
    if not pending:
        flash('Không tìm thấy thông tin đăng ký. Vui lòng đăng ký lại.', 'warning')
        return redirect(url_for('register'))

    if request.method == 'POST':
        if request.form.get('action') == 'resend':
            new_code = f"{random.randint(100000, 999999)}"
            session['pending_registration_code'] = new_code
            session['pending_registration_exp'] = (
                datetime.datetime.utcnow() + datetime.timedelta(minutes=10)
            ).isoformat()
            try:
                _send_verification_email(pending.get('ho_ten'), pending.get('email'), new_code)
                flash('Đã gửi lại mã xác nhận.', 'info')
            except Exception as e:
                flash(f'Không thể gửi lại email: {e}', 'danger')
            return redirect(url_for('verify_email'))

        code = request.form.get('verification_code', '').strip()
        stored_code = session.get('pending_registration_code')
        expiry_str = session.get('pending_registration_exp')
        expiry = datetime.datetime.fromisoformat(expiry_str) if expiry_str else None

        if expiry and datetime.datetime.utcnow() > expiry:
            flash('Mã xác nhận đã hết hạn. Vui lòng yêu cầu mã mới.', 'warning')
            return redirect(url_for('verify_email'))

        if not code or code != stored_code:
            flash('Mã xác nhận không chính xác.', 'danger')
            return redirect(url_for('verify_email'))

        try:
            conn = get_connection()
            cur = conn.cursor()
            cur.execute(
                """
                INSERT INTO NguoiDung (HoTen, GioiTinh, NgaySinh, Email, MatKhau, Role)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    pending['ho_ten'],
                    pending['gioi_tinh'],
                    pending['ngay_sinh'],
                    pending['email'],
                    pending['mat_khau'],
                    pending['role']
                )
            )
            conn.commit()
        except Exception as e:
            conn.rollback()
            conn.close()
            flash(f'Lỗi khi tạo tài khoản: {e}', 'danger')
            return redirect(url_for('register'))
        else:
            conn.close()
            _clear_pending_registration()
            flash('Đăng ký thành công! Vui lòng đăng nhập.', 'success')
            return redirect(url_for('login'))

    email = pending.get('email')
    return render_template('verify_email.html', email=email)


@app.route('/confirm_patient_invite/<token>', methods=['GET', 'POST'])
def confirm_patient_invite(token):
    try:
        data = decode_invite_token(token)
    except SignatureExpired:
        flash("Liên kết xác nhận đã hết hạn. Vui lòng liên hệ bác sĩ để gửi lại.", "warning")
        return redirect(url_for('login'))
    except BadSignature:
        flash("Liên kết xác nhận không hợp lệ.", "danger")
        return redirect(url_for('login'))

    if request.method == 'POST':
        code_input = request.form.get('verification_code', '').strip()
        if not code_input or code_input != data.get('code'):
            flash("Mã xác thực không chính xác.", "danger")
            return render_template('confirm_invite.html', email=data.get('email'), name=data.get('ho_ten'), token=token)

        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT ID FROM NguoiDung WHERE Email = ?", (data.get('email'),))
        if cur.fetchone():
            conn.close()
            flash("Tài khoản đã được kích hoạt trước đó. Bạn có thể đăng nhập.", "info")
            return redirect(url_for('login'))
        try:
            cur.execute("""
                INSERT INTO NguoiDung (HoTen, GioiTinh, NgaySinh, Email, MatKhau, DienThoai, DiaChi, Role)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data.get('ho_ten'),
                data.get('gioi_tinh'),
                data.get('ngay_sinh') or None,
                data.get('email'),
                data.get('mat_khau'),
                data.get('dien_thoai'),
                data.get('dia_chi'),
                data.get('role') or 'patient'
            ))
            conn.commit()
        except Exception as e:
            conn.rollback()
            flash(f"Lỗi khi tạo tài khoản: {e}", "danger")
            return render_template('confirm_invite.html', email=data.get('email'), name=data.get('ho_ten'), token=token)
        finally:
            conn.close()

        flash("Kích hoạt tài khoản thành công! Bạn có thể đăng nhập ngay bây giờ.", "success")
        session.clear()
        return redirect(url_for('login'))

    return render_template('confirm_invite.html', email=data.get('email'), name=data.get('ho_ten'), token=token)


@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = (request.form.get('email') or '').strip().lower()
        if not email:
            flash("Vui lòng nhập email đã đăng ký.", "warning")
            return render_template('forgot_password.html')

        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT ID, HoTen FROM NguoiDung WHERE Email = ?", (email,))
        user = cur.fetchone()
        if not user:
            conn.close()
            flash("Email không tồn tại trong hệ thống.", "danger")
            return render_template('forgot_password.html')

        new_pass = ''.join(random.choices(string.ascii_letters + string.digits, k=10))
        try:
            cur.execute("UPDATE NguoiDung SET MatKhau=? WHERE ID=?", (new_pass, user.ID))
            conn.commit()
        except Exception as e:
            conn.rollback()
            conn.close()
            flash(f"Lỗi khi cập nhật mật khẩu: {e}", "danger")
            return render_template('forgot_password.html')
        conn.close()

        email_body = f"""
        <div style='font-family:Arial,sans-serif;line-height:1.6'>
          <p>Xin chào {user.HoTen},</p>
          <p>Mật khẩu đăng nhập mới của bạn là:
            <strong style='font-size:1.2rem;'>{new_pass}</strong></p>
          <p>Vui lòng đăng nhập và đổi mật khẩu ngay để đảm bảo an toàn.</p>
        </div>
        """
        try:
            send_email(email, "Mật khẩu mới CVD-App", email_body)
            flash("Mật khẩu mới đã được gửi tới email của bạn.", "success")
            return redirect(url_for('login'))
        except Exception as e:
            flash(f"Không thể gửi email: {e}", "danger")
            return render_template('forgot_password.html')

    return render_template('forgot_password.html')

@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('username', '').strip().lower()
        pw = request.form.get('password', '').strip()

        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT ID, HoTen, Role, MatKhau
            FROM NguoiDung
            WHERE Email = ?
        """, (email,))
        user = cur.fetchone()
        conn.close()

        # 🔹 Kiểm tra tài khoản & mật khẩu
        if user and user.MatKhau == pw:
            # Tạo session
            session['user_id'] = user.ID
            session['user'] = user.HoTen
            session['role'] = user.Role

            # Hiển thị thông báo chào mừng
            flash(f"🎉 Chào mừng {user.HoTen} đăng nhập thành công!", "success")

            # ✅ Điều hướng theo vai trò
            if user.Role == 'admin':
                return redirect(url_for('history'))
            elif user.Role == 'doctor':
                return redirect(url_for('home'))  
            else:
                return redirect(url_for('home'))

        else:
            # ❌ Sai mật khẩu → hiển thị ngay
            flash("❌ Sai tài khoản hoặc mật khẩu. Vui lòng thử lại!", "danger")
            return render_template('login.html')

    # GET request → hiển thị form
    return render_template('login.html')

@app.route('/auth/<provider>')
def oauth_login(provider):
    provider = provider.lower()
    if provider not in SOCIAL_PROVIDERS:
        abort(404)
    if not SOCIAL_PROVIDERS.get(provider):
        flash("Chức năng đăng nhập này chưa được cấu hình.", "warning")
        return redirect(url_for('login'))

    client = oauth.create_client(provider)
    if not client:
        flash("Không tìm thấy cấu hình cho nhà cung cấp đăng nhập.", "danger")
        return redirect(url_for('login'))

    redirect_uri = url_for('oauth_callback', provider=provider, _external=True)
    kwargs = {}
    if provider == "google":
        nonce = token_urlsafe(16)
        session['oauth_nonce'] = nonce
        kwargs['nonce'] = nonce
    return client.authorize_redirect(redirect_uri, **kwargs)

@app.route('/auth/callback/<provider>')
def oauth_callback(provider):
    provider = provider.lower()
    if provider not in SOCIAL_PROVIDERS or not SOCIAL_PROVIDERS.get(provider):
        flash("Nhà cung cấp đăng nhập chưa được kích hoạt.", "warning")
        return redirect(url_for('login'))

    client = oauth.create_client(provider)
    if not client:
        flash("Không thể khởi tạo nhà cung cấp đăng nhập.", "danger")
        return redirect(url_for('login'))

    try:
        token = client.authorize_access_token()
        nonce = session.pop('oauth_nonce', None)
        user_info = client.parse_id_token(token, nonce=nonce)
    except Exception as e:
        flash(f"Không thể xác thực: {e}", "danger")
        return redirect(url_for('login'))

    email = user_info.get("email")
    if not email:
        flash("Không đọc được email từ tài khoản của bạn. Vui lòng cho phép truy cập email.", "warning")
        return redirect(url_for('login'))
    full_name = user_info.get("name") or user_info.get("given_name") or email.split("@")[0]

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT ID, HoTen, Role, MatKhau FROM NguoiDung WHERE Email = ?", (email,))
    user = cur.fetchone()

    if user:
        user_id = user.ID
        ho_ten = user.HoTen or full_name
        role = user.Role or 'patient'
        has_password = bool((user.MatKhau or "").strip())
    else:
        cur.execute("""
            INSERT INTO NguoiDung (HoTen, GioiTinh, NgaySinh, Email, MatKhau, Role)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (full_name, 'Nam', None, email, DEFAULT_SOCIAL_PASSWORD, 'patient'))
        conn.commit()
        cur.execute("SELECT ID, HoTen, Role, MatKhau FROM NguoiDung WHERE Email = ?", (email,))
        user = cur.fetchone()
        user_id = user.ID
        ho_ten = user.HoTen or full_name
        role = user.Role or 'patient'
        has_password = False

    needs_password_email = not has_password

    if needs_password_email:
        cur.execute("UPDATE NguoiDung SET MatKhau=? WHERE ID=?", (DEFAULT_SOCIAL_PASSWORD, user_id))
        conn.commit()
        conn.close()
        login_link = url_for('login', _external=True)

        session['user_id'] = user_id
        session['user'] = ho_ten
        session['role'] = role

        email_body = f"""
        <div style="font-family:Arial,sans-serif;line-height:1.6">
          <h2 style="color:#0d6efd">Chào {ho_ten},</h2>
          <p>Bạn vừa đăng ký/đăng nhập bằng Google trên CVD-App.</p>
          <p>Mật khẩu đăng nhập tạm thời của bạn là:
            <strong style="font-size:1.2rem;">{DEFAULT_SOCIAL_PASSWORD}</strong></p>
          <p>Nhấn vào liên kết sau để mở trang đăng nhập:
            <a href="{login_link}" style="color:#0d6efd; font-weight:bold;">Quay lại đăng nhập</a>
          </p>
          <p>Sau khi vào hệ thống, nhớ đổi mật khẩu trong phần Hồ sơ cá nhân để đảm bảo an toàn.</p>
          <p>Trân trọng,<br/>Đội ngũ CVD-App</p>
        </div>
        """
        try:
            send_email(
                to_email=email,
                subject="Mật khẩu đăng nhập CVD-App",
                html_body=email_body
            )
        except Exception as e:
            print(f"[WARN] Không thể gửi email mật khẩu Google: {e}")

        return render_template('google_password_sent.html', email=email)

    conn.close()

    session['user_id'] = user_id
    session['user'] = ho_ten
    session['role'] = role
    flash(f"🎉 Chào mừng {ho_ten} đăng nhập thành công!", "success")

    if role == 'admin':
        return redirect(url_for('history'))
    return redirect(url_for('home'))


# ==========================================
# Trang chủ
# ==========================================
@app.route('/home')
def home():
    if 'user' not in session:
        return redirect(url_for('login'))
    return render_template('home.html')

# ==========================================
# 📡 API: Lấy thông tin bệnh nhân từ hồ sơ NguoiDung
# ==========================================
@app.route('/get_patient_info/<int:benhnhan_id>')
def get_patient_info(benhnhan_id):
    if 'user' not in session or session.get('role') != 'doctor':
        return jsonify({"error": "Unauthorized"}), 403

    conn = get_connection()
    cur = conn.cursor()

    # 🔹 Lấy trực tiếp tuổi và giới tính từ hồ sơ NguoiDung
    cur.execute("""
        SELECT 
            DATEDIFF(YEAR, NgaySinh, GETDATE()) AS Tuoi,
            GioiTinh
        FROM NguoiDung
        WHERE ID = ?
    """, (benhnhan_id,))

    row = cur.fetchone()
    conn.close()

    if row:
        return jsonify({
            "tuoi": row.Tuoi,
            "gioitinh": row.GioiTinh
        })
    else:
        return jsonify({"tuoi": None, "gioitinh": None})

# ==========================================
# 🩺 Chẩn đoán bệnh tim mạch + Giải thích SHAP
# ==========================================
@app.route('/diagnose', methods=['GET', 'POST'])
def diagnose():
    if 'user' not in session:
        return redirect(url_for('login'))

    conn = get_connection()
    cur = conn.cursor()
    benhnhans = []
    if session.get('role') == 'doctor':
        cur.execute("SELECT ID, HoTen FROM NguoiDung WHERE Role='patient'")
        benhnhans = [
            {"ID": r.ID, "MaBN": f"BN{r.ID:03}", "HoTen": r.HoTen}
            for r in cur.fetchall()
        ]

    # --- Biến khởi tạo ---
    result = None
    ai_advice = None
    file_result = None
    risk_percent = None
    risk_level = None
    shap_file = None
    results = []    
    threshold = float(request.form.get('threshold', 0.5))

    # ======================
    # 🔹 XỬ LÝ NHẬP LIỆU THỦ CÔNG
    # ======================
    if request.method == 'POST' and 'predict_form' in request.form:
        try:
            benhnhan_id = (
                int(request.form.get('benhnhan_id'))
                if session.get('role') == 'doctor'
                else session['user_id']
            )

            # --- Lấy dữ liệu nhập tay ---
            age = int(request.form.get('age'))
            gender_raw = request.form.get('gender')
            gender = 1 if gender_raw == 'Nam' else 0
            weight = float(request.form.get('weight'))
            height = float(request.form.get('height'))
            bmi = round(weight / ((height / 100) ** 2), 2)
            systolic = float(request.form.get('systolic'))
            diastolic = float(request.form.get('diastolic'))
            chol = int(request.form.get('cholesterol'))
            glucose = int(request.form.get('glucose'))
            smoking = 1 if request.form.get('smoking') == 'yes' else 0
            alcohol = 1 if request.form.get('alcohol') == 'yes' else 0
            exercise = 1 if request.form.get('exercise') == 'yes' else 0

            # --- Dự đoán bằng mô hình ---
            if xgb_model:
                X = np.array([[
                    age, gender, height, weight, systolic, diastolic,
                    chol, glucose, smoking, alcohol, exercise
                ]], dtype=float)
                prob = float(xgb_model.predict_proba(X)[0, 1])
                risk_percent = round(prob * 100, 1)
                risk_level = 'high' if prob >= threshold else 'low'
            else:
                score = 0
                if systolic > 140 or diastolic > 90: score += 1
                score += chol
                score += glucose
                if bmi > 30: score += 1
                if smoking: score += 1
                if alcohol: score += 1
                prob = score / 8
                risk_percent = round(prob * 100, 1)
                risk_level = 'high' if prob >= threshold else 'low'

            nguy_co_text = "Nguy cơ cao" if risk_level == 'high' else "Nguy cơ thấp"
            result = f"{nguy_co_text} - {risk_percent}%"

            # --- Sinh lời khuyên AI ---
            chol_label = {0: "Bình thường", 1: "Cao nhẹ", 2: "Cao"}
            gluc_label = {0: "Bình thường", 1: "Cao nhẹ", 2: "Cao"}

            prompt = f"""
            Bạn là bác sĩ tim mạch.
            Dữ liệu bệnh nhân:
            - Tuổi: {age}
            - Giới tính: {gender_raw}
            - BMI: {bmi}
            - Huyết áp: {systolic}/{diastolic}
            - Cholesterol: {chol_label.get(chol, 'Không rõ')}
            - Đường huyết: {gluc_label.get(glucose, 'Không rõ')}
            - Hút thuốc: {'Có' if smoking else 'Không'}
            - Uống rượu bia: {'Có' if alcohol else 'Không'}
            - Tập thể dục: {'Có' if exercise else 'Không'}

            Ngưỡng dự đoán: {threshold}.
            Hãy đưa ra lời khuyên ngắn gọn, dễ hiểu, phù hợp với tình trạng trên.
            """

            ai_advice_raw = get_ai_advice_cached(prompt)
            ai_advice = highlight_advice(ai_advice_raw)

            # --- Sinh biểu đồ SHAP ---
            if xgb_model:
                try:
                    explainer = shap.TreeExplainer(xgb_model)
                    shap_values = explainer.shap_values(X)
                    shap.summary_plot(
                        shap_values, X,
                        feature_names=[
                            'Tuổi', 'Giới tính', 'Chiều cao', 'Cân nặng', 'HATT', 'HATTr',
                            'Cholesterol', 'Đường huyết', 'Hút thuốc', 'Rượu bia', 'Tập thể dục'
                        ],
                        show=False
                    )
                    shap_dir = os.path.join(app.root_path, 'static', 'images')
                    os.makedirs(shap_dir, exist_ok=True)
                    shap_file = f"shap_{benhnhan_id}.png"
                    plt.tight_layout()
                    plt.savefig(os.path.join(shap_dir, shap_file), bbox_inches='tight')
                    plt.close()
                except Exception as e:
                    print(f"⚠️ Lỗi khi tạo biểu đồ SHAP: {e}")

            # --- Lưu kết quả vào CSDL ---
            chol_label = {0: "Bình thường", 1: "Cao nhẹ", 2: "Cao"}
            gluc_label = {0: "Bình thường", 1: "Cao nhẹ", 2: "Cao"}

            bacsi_id = session['user_id'] if session.get('role') == 'doctor' else None
            cur.execute("""
                INSERT INTO ChanDoan
                (BenhNhanID, BacSiID, Tuoi, GioiTinh, BMI, HuyetApTamThu, HuyetApTamTruong,
                Cholesterol, DuongHuyet, HutThuoc, UongCon, TapTheDuc,
                NguyCo, LoiKhuyen, NgayChanDoan)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, GETDATE())
            """, (benhnhan_id, bacsi_id, age, gender_raw, bmi, systolic, diastolic,
                    chol_label.get(chol), gluc_label.get(glucose),
                    smoking, alcohol, exercise, nguy_co_text, ai_advice))
            conn.commit()

        except Exception as e:
            flash(f"Lỗi nhập liệu: {e}", "danger")

        # ======================
        # 🔹 XỬ LÝ FILE CSV / EXCEL
        # ======================
    if request.method == 'POST' and 'data_file' in request.files:
        try:
            file = request.files['data_file']
            if not file:
                flash("⚠️ Vui lòng chọn file CSV hoặc Excel trước khi tải lên.", "warning")
                return redirect(url_for('diagnose'))

            filename = file.filename.lower()
            if not filename.endswith(('.csv', '.xls', '.xlsx')):
                flash("❌ Chỉ hỗ trợ định dạng CSV, XLS hoặc XLSX", "danger")
                return redirect(url_for('diagnose'))

            # Đọc file theo định dạng
            if filename.endswith('.csv'):
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)

            # Chuẩn hóa tên cột
            df.columns = [c.strip().lower() for c in df.columns]

            # Các cột bắt buộc
            required_cols = ['age', 'gender', 'ap_hi', 'ap_lo', 'cholesterol',
                            'gluc', 'smoke', 'alco', 'active', 'weight', 'height']
            missing = [c for c in required_cols if c not in df.columns]
            if missing:
                flash(f"⚠️ File thiếu các cột: {', '.join(missing)}", "danger")
                return redirect(url_for('diagnose'))

            # Tính BMI
            df['bmi'] = (df['weight'] / ((df['height'] / 100) ** 2)).round(2)

            results = []
            for _, row in df.iterrows():
                age = int(row['age'])
                gender_raw = row['gender']
                gender = 1 if str(gender_raw).strip().lower() in ['nam', 'male', '1'] else 0
                height = float(row['height'])
                weight = float(row['weight'])
                systolic = float(row['ap_hi'])
                diastolic = float(row['ap_lo'])
                chol = int(row['cholesterol'])
                gluc = int(row['gluc'])
                smoking = int(row['smoke'])
                alcohol = int(row['alco'])
                exercise = int(row['active'])
                bmi = float(row['bmi'])

                # Dự đoán
                if xgb_model:
                    X = np.array([[
                        age, gender, height, weight, systolic, diastolic,
                        chol, gluc, smoking, alcohol, exercise
                    ]], dtype=float)
                    prob = float(xgb_model.predict_proba(X)[0, 1])
                else:
                    prob = 0.5

                risk_percent = round(prob * 100, 1)
                risk_level = "Nguy cơ cao" if prob >= threshold else "Nguy cơ thấp"

                results.append({
                    "Tuổi": age,
                    "Giới tính": gender_raw,
                    "Huyết áp": f"{systolic}/{diastolic}",
                    "Cholesterol": chol,
                    "Đường huyết": gluc,
                    "BMI": bmi,
                    "Hút thuốc": "Có" if smoking else "Không",
                    "Rượu/Bia": "Có" if alcohol else "Không",
                    "Tập thể dục": "Có" if exercise else "Không",
                    "Nguy cơ": risk_level,
                    "Xác suất (%)": risk_percent
                })

            file_result = pd.DataFrame(results).to_html(
                index=False,
                classes="table table-hover table-striped text-center align-middle small shadow-sm rounded-3"
            )

            flash("✅ Dự đoán từ file CSV/Excel đã hoàn tất!", "success")

        except Exception as e:
            flash(f"❌ Lỗi khi xử lý file CSV/Excel: {e}", "danger")


    conn.close()
    has_result = bool(result or ai_advice or shap_file or file_result)
    return render_template(
        'diagnose.html',
        benhnhans=benhnhans,
        result=result,
        risk_percent=risk_percent,
        risk_level=risk_level,
        threshold=threshold,
        ai_advice=ai_advice,
        file_result=file_result,
        shap_file=shap_file ,
        results=results,
        has_result=has_result 
    )

@app.route('/send-diagnosis-email', methods=['POST'])
def send_diagnosis_email():
    if 'user' not in session:
        return redirect(url_for('login'))

    risk_percent = request.form.get('risk_percent')
    risk_level = request.form.get('risk_level', 'low')
    threshold = request.form.get('threshold', '0.5')
    ai_advice_plain = request.form.get('ai_advice_plain', '').strip()
    benhnhan_id = request.form.get('benhnhan_id') or session.get('user_id')

    if not risk_percent:
        flash("Chưa có kết quả chẩn đoán để gửi email.", "warning")
        return redirect(url_for('diagnose'))

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT HoTen, Email FROM NguoiDung WHERE ID = ?", (benhnhan_id,))
    patient = cur.fetchone()
    conn.close()

    if not patient or not patient.Email:
        flash("Không tìm thấy email người nhận.", "danger")
        return redirect(url_for('diagnose'))

    patient_name = patient.HoTen or "bạn"
    risk_text = "Nguy cơ cao" if risk_level == 'high' else "Nguy cơ thấp"
    try:
        html_body = render_template(
            'emails/diagnosis_email.html',
            name=patient_name,
            risk_percent=risk_percent,
            risk_text=risk_text,
            threshold=threshold,
            ai_advice=ai_advice_plain,
            doctor_name=session.get('user', 'CVD-App')
        )
        send_email(
            patient.Email,
            "Kết quả chẩn đoán tim mạch từ CVD-App",
            html_body
        )
        flash("Đã gửi kết quả qua email. Cảm ơn bạn đã sử dụng dịch vụ!", "success")
    except Exception as e:
        flash(f"Không thể gửi email: {e}", "danger")

    return redirect(url_for('diagnose'))

# ==========================================
# 🎨 Hàm tô đậm lời khuyên AI (1 màu nhấn - FIX BUG "600;'>")
# ==========================================
import re

def highlight_advice(text):
    """💡 Làm nổi bật ý chính trong lời khuyên AI chỉ với 1 màu nhấn, an toàn không lỗi HTML."""
    if not text:
        return ""

    # Xóa ký tự markdown (** hoặc *)
    text = re.sub(r'\*{1,3}', '', text)

    # 🔹 Nhấn mạnh từ khóa (tích cực hoặc cảnh báo)
    keywords = [
        r"(hãy|nên|cần|duy trì|giữ|kiểm soát|theo dõi|tránh|không nên|quan trọng|nguy cơ|cao|béo phì|hút thuốc|rượu|bia|ngủ đủ|tập luyện|ăn uống|điều chỉnh)"
    ]

    for kw in keywords:
        text = re.sub(
            kw,
            lambda m: f"<b class='text-primary fw-semibold'>{m.group(0)}</b>",
            text,
            flags=re.IGNORECASE
        )

    # 🔹 Làm nổi bật các con số / phần trăm / đơn vị đo
    text = re.sub(
        r"\b\d+(\.\d+)?\s*(%|mmHg|kg|cm)?\b",
        lambda m: f"<b class='text-primary'>{m.group(0)}</b>",
        text
    )

    # 🔹 Thay newline bằng <br> cho trình bày đẹp
    text = re.sub(r'\n+', '<br>', text.strip())

    # 🔹 Gói khối nội dung
    text = f"""
    <div style="
        text-align: justify;
        line-height: 1.8;
        font-size: 15px;
        color: #212529;
    ">
        {text}
    </div>
    """

    return text

# ==========================================
# 📜 Lịch sử chẩn đoán (phân quyền + lọc bệnh nhân cho bác sỹ)
# ==========================================
@app.route('/history')
def history():
    if 'user' not in session:
        return redirect(url_for('login'))

    conn = get_connection()
    cur = conn.cursor()

    # ===== Lấy các tham số lọc từ URL =====
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    patient_id = request.args.get('patient_id')
    doctor_id = request.args.get('doctor_id')
    risk_filter = request.args.get('risk_filter')
    sort_order = request.args.get('sort', 'desc')

    # ===== Điều kiện mặc định =====
    where_clause = "WHERE 1=1"
    params = []

    # ===== Phân quyền =====
    role = session.get('role')

    if role == 'doctor':
        # 👨‍⚕️ Bác sỹ xem các ca do mình chẩn đoán
        where_clause += " AND BacSiID = ?"
        params.append(session['user_id'])
        # Và có thể lọc thêm theo bệnh nhân
        if patient_id:
            where_clause += " AND BenhNhanID = ?"
            params.append(patient_id)

    elif role == 'patient':
        # 🧑‍🦽 Bệnh nhân xem toàn bộ các ca của mình
        where_clause += " AND BenhNhanID = ?"
        params.append(session['user_id'])

    else:
        # 🧑‍💻 Admin xem toàn bộ, có thể lọc theo bác sỹ hoặc bệnh nhân
        if doctor_id:
            where_clause += " AND BacSiID = ?"
            params.append(doctor_id)
        if patient_id:
            where_clause += " AND BenhNhanID = ?"
            params.append(patient_id)

    # ===== Lọc theo ngày =====
    if start_date:
        where_clause += " AND NgayChanDoan >= CONVERT(DATE, ?)"
        params.append(start_date)
    if end_date:
        where_clause += " AND NgayChanDoan <= CONVERT(DATE, ?)"
        params.append(end_date)

    # ===== Lọc theo nguy cơ =====
    if risk_filter == 'high':
        where_clause += " AND LOWER(NguyCo) LIKE '%cao%'"
    elif risk_filter == 'low':
        where_clause += " AND LOWER(NguyCo COLLATE SQL_Latin1_General_Cp1253_CI_AI) LIKE '%thap%'"

    # ===== Truy vấn chính =====
    query = f"""
        SELECT ChanDoanID, BenhNhanID, TenBenhNhan, GioiTinh, Tuoi, TenBacSi, NgayChanDoan,
        BMI, HuyetApTamThu, HuyetApTamTruong, Cholesterol, DuongHuyet,
        HutThuoc, UongCon, TapTheDuc, NguyCo, LoiKhuyen

        FROM V_LichSuChanDoan
        {where_clause}
        ORDER BY NgayChanDoan {'DESC' if sort_order == 'desc' else 'ASC'}
    """

    cur.execute(query, params)
    records = cur.fetchall()
    conn.close()

    # ✅ Đếm tổng số bản ghi
    total_records = len(records)

    # ✅ Highlight lời khuyên
    try:
        from app import highlight_advice
        for r in records:
            if hasattr(r, "LoiKhuyen") and r.LoiKhuyen:
                r.LoiKhuyen = highlight_advice(r.LoiKhuyen)
    except Exception as e:
        print(f"⚠️ Lỗi highlight: {e}")

    # ===== Danh sách lọc =====
    doctors, patients = [], []

    if role == 'doctor':
        # Danh sách bệnh nhân mà bác sỹ đó đã chẩn đoán
        conn2 = get_connection()
        cur2 = conn2.cursor()
        cur2.execute("""
            SELECT DISTINCT bn.ID, bn.HoTen 
            FROM ChanDoan cd 
            JOIN NguoiDung bn ON cd.BenhNhanID = bn.ID
            WHERE cd.BacSiID = ?
        """, (session['user_id'],))
        patients = cur2.fetchall()
        conn2.close()

    elif role == 'admin':
        # Danh sách bác sỹ và bệnh nhân cho admin
        conn2 = get_connection()
        cur2 = conn2.cursor()
        cur2.execute("SELECT ID, HoTen FROM NguoiDung WHERE Role='doctor'")
        doctors = cur2.fetchall()
        cur2.execute("SELECT ID, HoTen FROM NguoiDung WHERE Role='patient'")
        patients = cur2.fetchall()
        conn2.close()

    # ===== Render =====
    return render_template(
        'history.html',
        records=records,
        doctors=doctors,
        patients=patients,
        start_date=start_date,
        end_date=end_date,
        patient_id=patient_id,
        doctor_id=doctor_id,
        risk_filter=risk_filter,
        sort_order=sort_order,
        total_records=total_records
    )


# ==========================================
# 🗑️ Xóa bản ghi chẩn đoán
# ==========================================
@app.route('/delete_history/<int:id>', methods=['POST'])
def delete_history(id):
    if 'user' not in session:
        return redirect(url_for('login'))

    role = session.get('role')
    if role not in ['doctor', 'admin','patient']:
        flash("❌ Bạn không có quyền xóa bản ghi chẩn đoán.", "danger")
        return redirect(url_for('history'))

    conn = get_connection()
    cur = conn.cursor()
    try:
        # ✅ Xóa theo ID (khóa chính)
        cur.execute("DELETE FROM ChanDoan WHERE ID = ?", (id,))
        conn.commit()
        flash("🗑️ Đã xóa bản ghi chẩn đoán thành công!", "success")

    except Exception as e:
        conn.rollback()
        flash(f"❌ Lỗi khi xóa bản ghi: {e}", "danger")

    finally:
        conn.close()

    return redirect(url_for('history'))

# ==========================================
# Chỉnh sửa lời khuyên (chỉ dành cho bác sỹ)
# ==========================================
@app.route('/edit_advice/<int:id>', methods=['POST'])
def edit_advice(id):
    if 'user' not in session or session.get('role') != 'doctor':
        flash("❌ Bạn không có quyền chỉnh sửa lời khuyên.", "danger")
        return redirect(url_for('login'))

    new_advice = request.form.get('loi_khuyen', '').strip()

    # 🧹 Làm sạch: loại bỏ mọi thẻ HTML, style còn sót lại
    import re
    from html import unescape
    clean_text = re.sub(r'<[^>]+>', '', new_advice)    # xóa thẻ HTML
    clean_text = unescape(clean_text)                # giải mã HTML entities (&nbsp;)
    clean_text = re.sub(r'\s{2,}', ' ', clean_text)   # gộp khoảng trắng

    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            UPDATE ChanDoan
            SET LoiKhuyen = ?
            WHERE ID = ?
        """, (clean_text, id))
        conn.commit()
        flash("✅ Đã cập nhật lời khuyên cho bệnh nhân.", "success")

    except Exception as e:
        conn.rollback()
        flash(f"❌ Lỗi khi cập nhật lời khuyên: {e}", "danger")

    finally:
        conn.close()

    return redirect(url_for('history'))


# ==========================================
# Quản lý tài khoản & hồ sơ bệnh nhân (phiên bản giới hạn quyền)
# ==========================================
@app.route('/manage_accounts', methods=['GET', 'POST'])
def manage_accounts():
    # ✅ Chỉ cho phép bác sỹ truy cập
    if 'user' not in session or session.get('role') != 'doctor':
        return redirect(url_for('login'))

    conn = get_connection()
    cur = conn.cursor()

    # ================================
    # ➕ THÊM bệnh nhân mới
    # ================================
    if request.method == 'POST' and 'add_patient' in request.form:
        ho_ten = request.form.get('ho_ten')
        gioi_tinh = request.form.get('gioi_tinh')
        ngay_sinh = request.form.get('ngay_sinh')
        email = (request.form.get('email') or '').strip().lower()
        mat_khau = request.form.get('mat_khau')
        dien_thoai = request.form.get('dien_thoai')
        dia_chi = request.form.get('dia_chi')

        if not is_valid_gmail(email):
            flash("Vui lòng nhập Gmail hợp lệ (ví dụ: ten@gmail.com).", "warning")
        else:
            cur.execute("SELECT COUNT(*) FROM NguoiDung WHERE Email = ?", (email,))
            if cur.fetchone()[0] > 0:
                flash("Email này đã tồn tại trong hệ thống.", "warning")
            else:
                verification_code = f"{random.randint(100000, 999999)}"
                payload = {
                    "ho_ten": ho_ten,
                    "gioi_tinh": gioi_tinh,
                    "ngay_sinh": ngay_sinh,
                    "email": email,
                    "mat_khau": mat_khau,
                    "dien_thoai": dien_thoai,
                    "dia_chi": dia_chi,
                    "role": "patient",
                    "code": verification_code
                }
                token = generate_invite_token(payload)
                confirm_link = url_for('confirm_patient_invite', token=token, _external=True)
                email_body = f"""
                <div style='font-family:Arial,sans-serif;line-height:1.6'>
                  <h2 style='color:#0d6efd'>Chào {ho_ten}</h2>
                  <p>Bạn được bác sĩ thêm vào hệ thống CVD-App. Vui lòng xác nhận email để kích hoạt tài khoản.</p>
                  <p>Mã xác thực của bạn: <strong style='font-size:1.2rem;'>{verification_code}</strong></p>
                  <p>Nhấn vào liên kết sau để hoàn tất: <a href='{confirm_link}'>Xác nhận tài khoản</a></p>
                </div>
                """
                try:
                    send_email(email, "Xác nhận tài khoản CVD-App", email_body)
                    flash("Đã gửi email xác thực tới bệnh nhân. Vui lòng yêu cầu bệnh nhân kiểm tra hộp thư.", "success")
                except Exception as e:
                    flash(f"Không thể gửi email xác thực: {e}", "danger")

    # ================================
    # 🗑️ XÓA tài khoản bệnh nhân (chỉ nếu bác sỹ từng chẩn đoán)
    # ================================
    if request.method == 'POST' and 'delete_patient' in request.form:
        patient_id = int(request.form.get('id'))
        doctor_id = session['user_id']

        # Kiểm tra quyền trước khi xóa
        cur.execute("""
            SELECT COUNT(*) FROM ChanDoan 
            WHERE BacSiID=? AND BenhNhanID=?
        """, (doctor_id, patient_id))
        has_permission = cur.fetchone()[0] > 0

        if not has_permission:
            flash("🚫 Bạn không có quyền xóa bệnh nhân này (chưa từng chẩn đoán).", "danger")
        else:
            try:
                cur.execute("DELETE FROM ChanDoan WHERE BenhNhanID=?", (patient_id,))
                cur.execute("DELETE FROM TinNhanAI WHERE BenhNhanID=?", (patient_id,))
                cur.execute("DELETE FROM NguoiDung WHERE ID=?", (patient_id,))
                conn.commit()
                flash("🗑️ Đã xóa tài khoản và toàn bộ lịch sử chẩn đoán của bệnh nhân.", "success")
            except Exception as e:
                conn.rollback()
                flash(f"❌ Lỗi khi xóa: {e}", "danger")

    # ================================
    # ✍️ CẬP NHẬT thông tin bệnh nhân (chỉ nếu bác sỹ từng chẩn đoán)
    # ================================
    if request.method == 'POST' and 'update_patient' in request.form:
        patient_id = int(request.form.get('id'))
        doctor_id = session['user_id']

        cur.execute("""
            SELECT COUNT(*) FROM ChanDoan 
            WHERE BacSiID=? AND BenhNhanID=?
        """, (doctor_id, patient_id))
        has_permission = cur.fetchone()[0] > 0

        if not has_permission:
            flash("🚫 Bạn không có quyền chỉnh sửa bệnh nhân này (chưa từng chẩn đoán).", "danger")
        else:
            try:
                cur.execute("""
                    UPDATE NguoiDung
                    SET HoTen=?, GioiTinh=?, NgaySinh=?, DienThoai=?, DiaChi=?
                    WHERE ID=?
                """, (
                    request.form.get('ho_ten'),
                    request.form.get('gioi_tinh'),
                    request.form.get('ngay_sinh'),
                    request.form.get('dien_thoai'),
                    request.form.get('dia_chi'),
                    patient_id
                ))
                conn.commit()
                flash("✅ Đã cập nhật thông tin bệnh nhân.", "success")
            except Exception as e:
                conn.rollback()
                flash(f"❌ Lỗi khi cập nhật: {e}", "danger")

    # ================================
    # 🔎 TÌM KIẾM bệnh nhân
    # ================================
    search = request.args.get('search', '').strip()

    if search:
        cur.execute("""
            SELECT ID, HoTen, Email, GioiTinh, NgaySinh, DienThoai, DiaChi
            FROM NguoiDung
            WHERE Role='patient' AND (HoTen LIKE ? OR Email LIKE ?)
            ORDER BY HoTen
        """, (f"%{search}%", f"%{search}%"))
    else:
        cur.execute("""
            SELECT ID, HoTen, Email, GioiTinh, NgaySinh, DienThoai, DiaChi
            FROM NguoiDung
            WHERE Role='patient'
            ORDER BY HoTen
        """)

    raw_patients = cur.fetchall()

    # ================================
    # 🔑 Lấy danh sách bệnh nhân bác sỹ từng chẩn đoán
    # ================================
    cur.execute("""
        SELECT DISTINCT BenhNhanID FROM ChanDoan WHERE BacSiID=?
    """, (session['user_id'],))
    my_patients = {r.BenhNhanID for r in cur.fetchall()}

    # ================================
    # XỬ LÝ dữ liệu hiển thị
    # ================================
    patients = []
    for p in raw_patients:
        if p.NgaySinh and hasattr(p.NgaySinh, "strftime"):
            ngay_sinh_str = p.NgaySinh.strftime("%d/%m/%Y")
            ngay_sinh_val = p.NgaySinh.strftime("%Y-%m-%d")
        else:
            ngay_sinh_str = p.NgaySinh if p.NgaySinh else "—"
            ngay_sinh_val = p.NgaySinh if p.NgaySinh else ""

        patients.append({
            "ID": p.ID,
            "HoTen": p.HoTen,
            "Email": p.Email,
            "GioiTinh": p.GioiTinh,
            "NgaySinh_str": ngay_sinh_str,
            "NgaySinh_val": ngay_sinh_val,
            "DienThoai": p.DienThoai,
            "DiaChi": p.DiaChi
        })

    conn.close()

    # ✅ Truyền thêm danh sách quyền my_patients sang template
    return render_template(
        'manage_accounts.html',
        patients=patients,
        search=search,
        my_patients=my_patients
    )

from flask import flash
from werkzeug.security import check_password_hash, generate_password_hash

import re
from flask import jsonify

# ==========================================
# 🔐 Đổi mật khẩu (xử lý AJAX)
# ==========================================
@app.route('/change_password', methods=['POST'])
def change_password():
    if 'user' not in session:
        return jsonify({"success": False, "message": "Vui lòng đăng nhập lại."}), 403

    old_pw = request.form.get('old_password')
    new_pw = request.form.get('new_password')
    confirm_pw = request.form.get('confirm_password')

    if not old_pw or not new_pw or not confirm_pw:
        return jsonify({"success": False, "message": "Vui lòng nhập đầy đủ thông tin."})

    if new_pw != confirm_pw:
        return jsonify({"success": False, "message": "Mật khẩu xác nhận không khớp."})

    # 🧹 Kiểm tra độ mạnh mật khẩu (ít nhất 8 ký tự, có hoa, số, đặc biệt)
    pattern = r'^(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&._\-#]).{8,}$'

    if not re.match(pattern, new_pw):
        return jsonify({
            "success": False,
            "message": "Mật khẩu phải ≥8 ký tự, chứa ít nhất 1 chữ hoa, 1 số và 1 ký tự đặc biệt."
        })

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT MatKhau FROM NguoiDung WHERE ID=?", (session['user_id'],))
    row = cur.fetchone()

    if not row or row.MatKhau != old_pw:
        conn.close()
        return jsonify({"success": False, "message": "Mật khẩu cũ không chính xác."})

    cur.execute("UPDATE NguoiDung SET MatKhau=? WHERE ID=?", (new_pw, session['user_id']))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "Đổi mật khẩu thành công!"})


# ==========================================
# Hồ sơ cá nhân
# ==========================================
@app.route('/profile', methods=['GET', 'POST'])
def profile():
    if 'user' not in session:
        return redirect(url_for('login'))

    conn = get_connection()
    cur = conn.cursor()

    # --- Khi người dùng cập nhật hồ sơ ---
    if request.method == 'POST':
        cur.execute("""
            UPDATE NguoiDung
            SET HoTen=?, DienThoai=?, NgaySinh=?, GioiTinh=?, DiaChi=?
            WHERE ID=?
        """, (
            request.form.get('ho_ten'),
            request.form.get('dien_thoai'),
            request.form.get('ngay_sinh'),
            request.form.get('gioi_tinh'),
            request.form.get('dia_chi'),
            session['user_id']
        ))
        conn.commit()

        # Lưu thời gian cập nhật tạm vào session
        from datetime import datetime
        update_time = datetime.now().strftime("%d/%m/%Y %H:%M")
        if 'timeline' not in session:
            session['timeline'] = []
        session['timeline'].insert(0, f"Cập nhật hồ sơ - {update_time}")

        flash("Cập nhật hồ sơ thành công!", "success")

    # --- Lấy thông tin người dùng (bao gồm ngày tạo tài khoản) ---
    cur.execute("""
        SELECT HoTen, Email, Role, DienThoai, NgaySinh, GioiTinh, DiaChi, MatKhau, NgayTao
        FROM NguoiDung WHERE ID=?
    """, (session['user_id'],))
    user_info = cur.fetchone()
    can_change_password = False
    if user_info:
        mat_khau_val = getattr(user_info, 'MatKhau', None)
        if isinstance(mat_khau_val, str):
            mat_khau_val = mat_khau_val.strip()
        can_change_password = bool(mat_khau_val)
    conn.close()

    # --- Chuẩn bị timeline hiển thị ---
    timeline = []
    if user_info and user_info[-1]:
        # user_info[-1] = NgayTao
        created_at = user_info[-1].strftime("%d/%m/%Y %H:%M")
        timeline.append(f"Tạo tài khoản - {created_at}")
    if 'timeline' in session:
        timeline = session['timeline'] + timeline

    return render_template(
        'profile.html',
        user_info=user_info,
        user_timeline=timeline,
        can_change_password=can_change_password
    )

# ==========================================
# 📨 Xuất báo cáo kết quả chẩn đoán ra Excel 
# ==========================================
@app.route('/export_diagnosis', methods=['POST'])
def export_diagnosis():
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    from flask import send_file
    from datetime import datetime
    import os, re

    # ===== Dữ liệu từ form =====
    data = {key: request.form.get(key, '') for key in [
        'age', 'gender', 'bmi', 'systolic', 'diastolic', 'cholesterol',
        'glucose', 'smoking', 'alcohol', 'exercise',
        'risk_percent', 'risk_level', 'ai_advice', 'shap_file', 'benhnhan_id'
    ]}

    # ===== Lấy tên người đăng nhập & vai trò =====
    user_name = session.get('user', 'Người dùng')
    user_role = session.get('role', 'patient')

    # ===== Xác định tên bệnh nhân và bác sỹ =====
    patient_name = None
    doctor_name = None

    if user_role == 'doctor':
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT HoTen FROM NguoiDung WHERE ID = ?", data.get('benhnhan_id'))
        row = cur.fetchone()
        conn.close()
        patient_name = row[0] if row else "Không xác định"
        doctor_name = user_name
    else:
        patient_name = user_name
        doctor_name = "—"

    # ===== Tạo workbook =====
    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo chẩn đoán"

    # ===== Style =====
    title_font = Font(size=18, bold=True, color="1F4E78")
    header_font = Font(size=13, bold=True, color="FFFFFF")
    section_font = Font(size=12, bold=True, color="1F4E78")
    normal_font = Font(size=11)
    advice_font = Font(size=12, color="000000")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_sub = PatternFill(start_color="E9F3FF", end_color="E9F3FF", fill_type="solid")
    fill_high = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    fill_low = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")

    # ===== Tiêu đề =====
    ws.merge_cells("A1:E1")
    ws["A1"] = "BÁO CÁO KẾT QUẢ CHẨN ĐOÁN TIM MẠCH"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws.append([])

    # ===== I. Thông tin chung =====
    ws.merge_cells("A3:E3")
    ws["A3"] = "I. THÔNG TIN CHUNG"
    ws["A3"].font = section_font
    ws["A3"].alignment = left

    ws.append(["Tên bệnh nhân", patient_name])
    ws.append(["Bác sỹ chẩn đoán", doctor_name])
    ws.append(["Ngày tạo báo cáo", datetime.now().strftime("%d/%m/%Y %H:%M")])
    ws.append([])

    # ===== II. Dữ liệu đầu vào =====
    ws.merge_cells("A7:E7")
    ws["A7"] = "II. DỮ LIỆU ĐẦU VÀO"
    ws["A7"].font = section_font
    ws["A7"].alignment = left

    ws.append(["Thuộc tính", "Giá trị", "Thuộc tính", "Giá trị"])
    for cell in ws[8]:
        cell.font = header_font
        cell.fill = fill_header
        cell.border = border
        cell.alignment = center

    input_data = [
        ["Tuổi", data['age'], "Giới tính", data['gender']],
        ["BMI", data['bmi'], "Huyết áp (HATT/HATTr)", f"{data['systolic']}/{data['diastolic']}"],
        ["Cholesterol", data['cholesterol'], "Đường huyết", data['glucose']],
        ["Hút thuốc", "Có" if data['smoking']=="yes" else "Không", "Rượu/Bia", "Có" if data['alcohol']=="yes" else "Không"],
        ["Tập thể dục", "Có" if data['exercise']=="yes" else "Không", "", ""]
    ]
    for row in input_data:
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.font = normal_font
            cell.border = border
            cell.alignment = left

    ws.append([])

        # ===== III. Kết quả chẩn đoán =====
    ws.merge_cells(f"A{ws.max_row+1}:E{ws.max_row+1}")
    ws[f"A{ws.max_row}"] = "III. KẾT QUẢ CHẨN ĐOÁN"
    ws[f"A{ws.max_row}"].font = section_font
    ws[f"A{ws.max_row}"].alignment = left

    ws.append(["Nguy cơ", "Tỷ lệ (%)", "Đánh giá", ""])
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = fill_header
        cell.border = border
        cell.alignment = center

    ws.append([
        "Cao" if data['risk_level'] == 'high' else "Thấp",
        data['risk_percent'] + "%",
        "⚠️ Cần theo dõi" if data['risk_level'] == 'high' else "✅ Ổn định",
        ""
    ])
    for cell in ws[ws.max_row]:
        cell.font = normal_font
        cell.border = border
        cell.alignment = center
        cell.fill = fill_high if data['risk_level'] == 'high' else fill_low

    ws.append([])

    import re
    from html import unescape

    # ===== Làm sạch và định dạng lời khuyên =====
    advice_raw = data.get('ai_advice') or "Chưa có lời khuyên từ AI."

    # ✅ Bỏ toàn bộ thẻ HTML & thuộc tính style
    advice_text = re.sub(r'style="[^"]*"', '', advice_raw)    # xóa thuộc tính style
    advice_text = re.sub(r'<[^>]+>', '', advice_text)         # xóa thẻ HTML còn lại
    advice_text = unescape(advice_text)                      # giải mã HTML entity (&nbsp;,...)
    advice_text = re.sub(r'\s*\n\s*', '\n', advice_text.strip())
    advice_text = re.sub(r'\s{2,}', ' ', advice_text)

    # ✅ Tự động ngắt dòng sau dấu chấm (khi sau đó là chữ in hoa hoặc tiếng Việt có dấu)
    advice_text = re.sub(r'\.\s*(?=[A-ZÀ-ỹ])', '.\n', advice_text)

    # ✅ Ngắt dòng trước các cụm từ như “Lời khuyên”, “Khuyến nghị”, “Tóm lại”
    advice_text = re.sub(r'(?=\b(Lời khuyên|Khuyến nghị|Tóm lại)\b)', '\n', advice_text)

    # ✅ Loại bỏ dòng trống dư
    advice_text = re.sub(r'\n{2,}', '\n', advice_text).strip()

    # ✅ Ghi ra Excel (xuống dòng, căn đều 2 bên)
    start_row = ws.max_row + 1
    end_row = start_row + 8
    ws.merge_cells(f"A{start_row}:E{end_row}")
    cell = ws[f"A{start_row}"]
    cell.value = advice_text
    cell.alignment = Alignment(horizontal="justify", vertical="top", wrap_text=True)
    cell.font = advice_font
    cell.border = border
    cell.fill = fill_sub

    # ===== V. Biểu đồ SHAP =====
    shap_path = os.path.join(app.root_path, 'static', 'images', data['shap_file']) if data['shap_file'] else None
    if shap_path and os.path.exists(shap_path):
        ws.merge_cells(f"A{ws.max_row+1}:E{ws.max_row+1}")
        ws[f"A{ws.max_row}"] = "V. GIẢI THÍCH KẾT QUẢ BẰNG BIỂU ĐỒ SHAP"
        ws[f"A{ws.max_row}"].font = section_font
        ws[f"A{ws.max_row}"].alignment = left
        try:
            img = ExcelImage(shap_path)
            img.width = 520
            img.height = 320
            ws.add_image(img, f"A{ws.max_row+1}")
        except Exception as e:
            ws.append([f"Lỗi khi chèn hình: {e}"])

    # ===== Footer =====
    ws.append([])
    ws.merge_cells(f"A{ws.max_row}:E{ws.max_row}")
    ws[f"A{ws.max_row}"] = f"🗓️ Báo cáo được tạo bởi: {doctor_name or user_name} — {datetime.now().strftime('%H:%M, %d/%m/%Y')}"
    ws[f"A{ws.max_row}"].alignment = center
    ws[f"A{ws.max_row}"].font = Font(size=10, italic=True, color="777777")

    # ===== Căn chỉnh độ rộng =====
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 10

    # ===== Xuất file =====
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"BaoCao_ChanDoan_{patient_name.replace(' ', '_')}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ==========================================
# 🚀 Đăng xuất
# ==========================================
@app.route('/logout')
def logout():
    if 'user' in session:
        name = session.get('user', 'Người dùng')
        session.clear()
        flash(f"👋 {name}, bạn đã đăng xuất khỏi hệ thống thành công!", "success")
    else:
        flash("⚠️ Bạn chưa đăng nhập!", "warning")
    return redirect(url_for('login'))

# =========================================================
# 📊 DASHBOARD THỐNG KÊ (Admin - Bản nâng cấp chuyên sâu)
# =========================================================
@app.route('/admin/dashboard')
def admin_dashboard():
    # --- Kiểm tra quyền truy cập ---
    if 'user' not in session or session.get('role') != 'admin':
        flash("Bạn không có quyền truy cập trang này!", "danger")
        return redirect(url_for('login'))

    conn = get_connection()
    cur = conn.cursor()

    # ========================== 1️⃣ TỔNG QUAN ==========================
    cur.execute("SELECT COUNT(*) FROM NguoiDung WHERE Role='doctor'")
    total_doctors = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM NguoiDung WHERE Role='patient'")
    total_patients = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM ChanDoan")
    total_diagnoses = cur.fetchone()[0]

    # ========================== 2️⃣ XU HƯỚNG CHẨN ĐOÁN ==========================
    cur.execute("""
        SELECT FORMAT(NgayChanDoan, 'MM-yyyy') AS Thang,
               COUNT(*) AS SoLuong,
               SUM(CASE WHEN LOWER(NguyCo) LIKE '%cao%' THEN 1 ELSE 0 END) AS SoCao
        FROM ChanDoan
        GROUP BY FORMAT(NgayChanDoan, 'MM-yyyy')
        ORDER BY MIN(NgayChanDoan)
    """)
    monthly = cur.fetchall()
    months = [row.Thang for row in monthly]
    counts = [row.SoLuong for row in monthly]
    high_risk = [row.SoCao for row in monthly]

    # ========================== 3️⃣ TỶ LỆ NGUY CƠ ==========================
    cur.execute("""
        SELECT NguyCo, COUNT(*) AS SoLuong
        FROM ChanDoan
        GROUP BY NguyCo
    """)
    risk_data = cur.fetchall()
    risk_labels = [row.NguyCo for row in risk_data]
    risk_values = [row.SoLuong for row in risk_data]

    # ========================== 4️⃣ TOP BÁC SỸ ==========================
    cur.execute("""
        SELECT TOP 5 bs.HoTen, COUNT(cd.ID) AS SoCa,
               SUM(CASE WHEN cd.NguyCo LIKE '%cao%' THEN 1 ELSE 0 END) * 100.0 / COUNT(*) AS TyLeCao
        FROM ChanDoan cd
        JOIN NguoiDung bs ON cd.BacSiID = bs.ID
        GROUP BY bs.HoTen
        ORDER BY SoCa DESC
    """)
    top_doctors = cur.fetchall()
    top_names = [row.HoTen for row in top_doctors]
    top_counts = [row.SoCa for row in top_doctors]
    top_rates = [round(row.TyLeCao, 1) for row in top_doctors]

    # ========================== 5️⃣ TRUNG BÌNH CHỈ SỐ Y KHOA ==========================
    cur.execute("""
        SELECT 
            ROUND(AVG(BMI), 1) AS AvgBMI,
            ROUND(AVG(HuyetApTamThu), 0) AS AvgHATT,
            ROUND(AVG(HuyetApTamTruong), 0) AS AvgHATTr,
            SUM(CASE WHEN HutThuoc=1 THEN 1 ELSE 0 END)*100.0/COUNT(*) AS SmokePercent,
            SUM(CASE WHEN UongCon=1 THEN 1 ELSE 0 END)*100.0/COUNT(*) AS AlcoPercent,
            SUM(CASE WHEN TapTheDuc=1 THEN 1 ELSE 0 END)*100.0/COUNT(*) AS ActivePercent
        FROM ChanDoan
    """)
    row = cur.fetchone()
    avg_bmi = row.AvgBMI or 0
    avg_systolic = row.AvgHATT or 0
    avg_diastolic = row.AvgHATTr or 0
    smoke_percent = round(row.SmokePercent or 0, 1)
    alco_percent = round(row.AlcoPercent or 0, 1)
    active_percent = round(row.ActivePercent or 0, 1)

    # ========================== 6️⃣ HIỆU SUẤT BÁC SỸ ==========================
    cur.execute("""
        SELECT ND.HoTen AS BacSi,
               COUNT(CD.ID) AS SoCa,
               SUM(CASE WHEN CD.NguyCo LIKE '%cao%' THEN 1 ELSE 0 END)*100.0/COUNT(*) AS TyLeCao
        FROM ChanDoan CD
        JOIN NguoiDung ND ON CD.BacSiID = ND.ID
        GROUP BY ND.HoTen
        ORDER BY SoCa DESC
    """)
    perf_rows = cur.fetchall()
    perf_names = [r.BacSi for r in perf_rows]
    perf_cases = [r.SoCa for r in perf_rows]
    perf_rate = [round(r.TyLeCao or 0, 1) for r in perf_rows]

    # ========================== 7️⃣ TỔNG SỐ BỆNH NHÂN CÓ CHẨN ĐOÁN ==========================
    cur.execute("SELECT COUNT(DISTINCT BenhNhanID) FROM ChanDoan")
    diagnosed_patients = cur.fetchone()[0]

    conn.close()

    return render_template(
        'admin_dashboard.html',
        total_doctors=total_doctors,
        total_patients=total_patients,
        total_diagnoses=total_diagnoses,
        diagnosed_patients=diagnosed_patients,
        months=months,
        counts=counts,
        high_risk=high_risk,
        risk_labels=risk_labels,
        risk_values=risk_values,
        top_names=top_names,
        top_counts=top_counts,
        top_rates=top_rates,
        avg_bmi=avg_bmi,
        avg_systolic=avg_systolic,
        avg_diastolic=avg_diastolic,
        smoke_percent=smoke_percent,
        alco_percent=alco_percent,
        active_percent=active_percent,
        perf_names=perf_names,
        perf_cases=perf_cases,
        perf_rate=perf_rate
    )

# =========================================================
# 🧑‍⚕️ Quản lý người dùng (Bác sỹ / Bệnh nhân) — Admin
# =========================================================
@app.route('/admin/manage_users', methods=['GET', 'POST'])
def admin_manage_users():
    if 'user' not in session or session.get('role') != 'admin':
        flash("❌ Bạn không có quyền truy cập trang này!", "danger")
        return redirect(url_for('login'))

    import datetime
    conn = get_connection()
    cur = conn.cursor()

    # Xác định loại người dùng đang quản lý
    role_type = request.args.get('type', 'doctor')  # mặc định là doctor
    title_map = {'doctor': 'Bác sỹ', 'patient': 'Bệnh nhân'}
    page_title = f"Quản lý {title_map.get(role_type, 'N/A')}"
    
    # ===================================================
    # ➕ THÊM NGƯỜI DÙNG
    # ===================================================
    if request.method == 'POST' and 'add_user' in request.form:
        ho_ten = request.form.get('ho_ten', '').strip()
        email = (request.form.get('email', '').strip().lower())
        mat_khau = request.form.get('mat_khau', '').strip()
        gioi_tinh = request.form.get('gioi_tinh')
        ngay_sinh = request.form.get('ngay_sinh') or None
        dien_thoai = request.form.get('dien_thoai')
        dia_chi = request.form.get('dia_chi')

        if not is_valid_gmail(email):
            flash("Vui lòng nhập Gmail hợp lệ (ví dụ: ten@gmail.com).", "warning")
        else:
            cur.execute("SELECT COUNT(*) FROM NguoiDung WHERE Email = ?", (email,))
            if cur.fetchone()[0] > 0:
                flash("Email này đã tồn tại trong hệ thống.", "warning")
            else:
                verification_code = f"{random.randint(100000, 999999)}"
                payload = {
                    "ho_ten": ho_ten,
                    "gioi_tinh": gioi_tinh,
                    "ngay_sinh": ngay_sinh,
                    "email": email,
                    "mat_khau": mat_khau,
                    "dien_thoai": dien_thoai,
                    "dia_chi": dia_chi,
                    "role": role_type,
                    "code": verification_code
                }
                token = generate_invite_token(payload)
                confirm_link = url_for('confirm_patient_invite', token=token, _external=True)
                email_body = f"""
                <div style='font-family:Arial,sans-serif;line-height:1.6'>
                  <h2 style='color:#0d6efd'>Chào {ho_ten}</h2>
                  <p>Bạn được quản trị viên thêm vào hệ thống CVD-App. Vui lòng xác nhận email để kích hoạt tài khoản.</p>
                  <p>Mã xác thực của bạn: <strong style='font-size:1.2rem;'>{verification_code}</strong></p>
                  <p>Nhấn vào liên kết sau để hoàn tất: <a href='{confirm_link}'>Xác nhận tài khoản</a></p>
                </div>
                """
                try:
                    send_email(email, "Xác nhận tài khoản CVD-App", email_body)
                    flash(f"Đã gửi email xác thực cho {title_map[role_type]}.", "success")
                except Exception as e:
                    flash(f"Không thể gửi email xác thực: {e}", "danger")

    # ===================================================
    # ✍️ SỬA NGƯỜI DÙNG
    # ===================================================
    elif request.method == 'POST' and 'edit_user' in request.form:
        id = request.form.get('id')
        ho_ten = request.form.get('ho_ten', '').strip()
        gioi_tinh = request.form.get('gioi_tinh')
        ngay_sinh = request.form.get('ngay_sinh') or None
        email = request.form.get('email', '').strip().lower()
        mat_khau = request.form.get('mat_khau', '').strip()
        dien_thoai = request.form.get('dien_thoai')
        dia_chi = request.form.get('dia_chi')

        if not mat_khau:
            cur.execute("""
                UPDATE NguoiDung
                SET HoTen=?, GioiTinh=?, NgaySinh=?, Email=?, DienThoai=?, DiaChi=?
                WHERE ID=? AND Role=?
            """, (ho_ten, gioi_tinh, ngay_sinh, email, dien_thoai, dia_chi, id, role_type))
        else:
            cur.execute("""
                UPDATE NguoiDung
                SET HoTen=?, GioiTinh=?, NgaySinh=?, Email=?, MatKhau=?, DienThoai=?, DiaChi=?
                WHERE ID=? AND Role=?
            """, (ho_ten, gioi_tinh, ngay_sinh, email, mat_khau, dien_thoai, dia_chi, id, role_type))
        conn.commit()
        flash(f"✍️ Cập nhật thông tin {title_map[role_type]} thành công!", "success")

    # ===================================================
    # 🗑️ XÓA NGƯỜI DÙNG
    # ===================================================
    elif request.method == 'POST' and 'delete_user' in request.form:
        user_id = int(request.form.get('id'))
        try:
            if role_type == 'patient':
                cur.execute("DELETE FROM ChanDoan WHERE BenhNhanID=?", (user_id,))
                cur.execute("DELETE FROM TinNhanAI WHERE BenhNhanID=?", (user_id,))
            elif role_type == 'doctor':
                # (Tùy chọn: Xử lý các ca chẩn đoán của bác sỹ này, ví dụ: set BacSiID = NULL)
                # cur.execute("UPDATE ChanDoan SET BacSiID=NULL WHERE BacSiID=?", (user_id,))
                pass # Hiện tại chỉ xóa bác sỹ

            cur.execute("DELETE FROM NguoiDung WHERE ID=? AND Role=?", (user_id, role_type))
            conn.commit()
            flash(f"Đã xóa {title_map[role_type]} khỏi hệ thống!", "success")
        except Exception as e:
            conn.rollback()
            flash(f"Lỗi khi xóa tài khoản: {e}", "danger")

    # ===================================================
    # 🧾 DANH SÁCH NGƯỜI DÙNG
    # ===================================================
    cur.execute(f"""
        SELECT ID, HoTen, Email, GioiTinh, NgaySinh, DienThoai, DiaChi, NgayTao
        FROM NguoiDung
        WHERE Role=?
        ORDER BY NgayTao DESC
    """, (role_type,))
    users = cur.fetchall()

    # Chuyển ngày sang kiểu datetime
    for u in users:
        if hasattr(u, 'NgaySinh') and isinstance(u.NgaySinh, str):
            try:
                u.NgaySinh = datetime.datetime.strptime(u.NgaySinh.split(" ")[0], "%Y-%m-%d")
            except:
                u.NgaySinh = None
        if hasattr(u, 'NgayTao') and isinstance(u.NgayTao, str):
            try:
                u.NgayTao = datetime.datetime.strptime(u.NgayTao.split(" ")[0], "%Y-%m-%d")
            except:
                u.NgayTao = None

    conn.close()

    return render_template('admin_users.html',
                            users=users,
                            role_type=role_type,
                            page_title=page_title)


# ==========================================
# 📊 XUẤT FILE EXCEL THỐNG KÊ HỆ THỐNG - Nâng cấp chuyên nghiệp
# ==========================================
@app.route('/export_admin_stats', methods=['POST'])
def export_admin_stats():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import PieChart, BarChart, LineChart, Reference
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from flask import send_file
    from datetime import datetime

    conn = get_connection()
    cur = conn.cursor()

    # =============================== #
    # 🗃️ LẤY DỮ LIỆU TỪ DATABASE
    # =============================== #
    cur.execute("SELECT COUNT(*) FROM NguoiDung WHERE Role='doctor'")
    total_doctors = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM NguoiDung WHERE Role='patient'")
    total_patients = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM ChanDoan")
    total_diagnoses = cur.fetchone()[0]

    cur.execute("""
        SELECT NguyCo, COUNT(*) AS SoLuong
        FROM ChanDoan
        GROUP BY NguyCo
    """)
    risk_data = cur.fetchall()

    cur.execute("""
        SELECT TOP 5 bs.HoTen, COUNT(cd.ID) AS SoCa
        FROM ChanDoan cd
        JOIN NguoiDung bs ON cd.BacSiID = bs.ID
        GROUP BY bs.HoTen
        ORDER BY SoCa DESC
    """)
    top_doctors = cur.fetchall()

    cur.execute("""
        SELECT ND.HoTen AS BacSi,
               COUNT(CD.ID) AS SoCa,
               SUM(CASE WHEN CD.NguyCo LIKE '%cao%' THEN 1 ELSE 0 END)*100.0/COUNT(*) AS TyLeCao
        FROM ChanDoan CD
        JOIN NguoiDung ND ON CD.BacSiID = ND.ID
        GROUP BY ND.HoTen
        ORDER BY SoCa DESC
    """)
    perf_rows = cur.fetchall()
    conn.close()

    # =============================== #
    # 📑 TẠO FILE EXCEL
    # =============================== #
    wb = Workbook()
    ws = wb.active
    ws.title = "Tổng quan hệ thống"

    # --- STYLE ---
    title_font = Font(size=16, bold=True, color="1F4E78")
    header_font = Font(size=12, bold=True, color="FFFFFF")
    fill_blue = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_gray = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999")
    )

    # =============================== #
    # 📈 SHEET 1: TỔNG QUAN
    # =============================== #
    ws.merge_cells("A1:E1")
    ws["A1"] = "BÁO CÁO THỐNG KÊ HỆ THỐNG CHẨN ĐOÁN TIM MẠCH"
    ws["A1"].font = title_font
    ws["A1"].alignment = align_center

    ws.append([])
    ws.append(["Ngày xuất báo cáo:", datetime.now().strftime("%d/%m/%Y %H:%M")])
    ws.append(["Người xuất:", session.get('user', 'Quản trị viên')])
    ws.append([])
    ws.append(["📊 Chỉ số tổng quan"])
    ws.append(["Tổng số bác sỹ", total_doctors])
    ws.append(["Tổng số bệnh nhân", total_patients])
    ws.append(["Tổng lượt chẩn đoán", total_diagnoses])
    ws.append([])

    ws.append(["🏆 Top 5 bác sỹ có số ca chẩn đoán nhiều nhất"])
    ws.append(["Tên bác sỹ", "Số ca"])
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = fill_blue
        cell.alignment = align_center
        cell.border = border

    for idx, d in enumerate(top_doctors, start=1):
        ws.append([d.HoTen, d.SoCa])
        for cell in ws[ws.max_row]:
            cell.border = border
            if idx % 2 == 0:
                cell.fill = fill_gray

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20

    # =============================== #
    # 📊 SHEET 2: TỶ LỆ BÁC SỸ / BỆNH NHÂN
    # =============================== #
    ws2 = wb.create_sheet("Bác sỹ - Bệnh nhân")
    ws2.append(["Loại tài khoản", "Số lượng"])
    ws2.append(["Bác sỹ", total_doctors])
    ws2.append(["Bệnh nhân", total_patients])

    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = fill_blue
        cell.alignment = align_center
        cell.border = border
    for row in ws2.iter_rows(min_row=2, max_col=2):
        for c in row:
            c.border = border
            c.alignment = align_center

    from openpyxl.chart.label import DataLabelList

    pie = PieChart()
    pie.title = "Tỷ lệ Bác sỹ / Bệnh nhân"
    data = Reference(ws2, min_col=2, min_row=1, max_row=3)
    labels = Reference(ws2, min_col=1, min_row=2, max_row=3)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)

    # ✅ Hiển thị giá trị + phần trăm + tên
    pie.dLbls = DataLabelList()
    pie.dLbls.showVal = True
    pie.dLbls.showPercent = True
    pie.dLbls.showCatName = True

    ws2.add_chart(pie, "D5")


    # =============================== #
    # 📊 SHEET 3: TỶ LỆ NGUY CƠ
    # =============================== #
    ws3 = wb.create_sheet("Nguy cơ cao - thấp")
    ws3.append(["Mức nguy cơ", "Số lượng"])
    for r in risk_data:
        ws3.append([r.NguyCo, r.SoLuong])

    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = fill_blue
        cell.alignment = align_center
        cell.border = border

    bar = BarChart()
    bar.title = "Tỷ lệ nguy cơ cao / thấp"
    data = Reference(ws3, min_col=2, min_row=1, max_row=ws3.max_row)
    cats = Reference(ws3, min_col=1, min_row=2, max_row=ws3.max_row)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.y_axis.title = "Số lượng"
    ws3.add_chart(bar, "E5")

    # =============================== #
    # 📊 SHEET 4: HIỆU SUẤT BÁC SỸ
    # =============================== #
    ws4 = wb.create_sheet("Hiệu suất bác sỹ")
    ws4.append(["Bác sỹ", "Số ca", "Tỷ lệ nguy cơ cao (%)"])
    for p in perf_rows:
        ws4.append([p.BacSi, p.SoCa, round(p.TyLeCao or 0, 1)])

    for cell in ws4[1]:
        cell.font = header_font
        cell.fill = fill_blue
        cell.alignment = align_center
        cell.border = border

    for row in ws4.iter_rows(min_row=2, max_col=3):
        for cell in row:
            cell.border = border
            if row[0].row % 2 == 0:
                cell.fill = fill_gray
            cell.alignment = align_center

    # --- Biểu đồ kết hợp ---
    chart = BarChart()
    chart.title = "Hiệu suất & Tỷ lệ nguy cơ cao của bác sỹ"
    chart.y_axis.title = "Số ca"
    data_bar = Reference(ws4, min_col=2, min_row=1, max_row=ws4.max_row)
    cats = Reference(ws4, min_col=1, min_row=2, max_row=ws4.max_row)
    chart.add_data(data_bar, titles_from_data=True)
    chart.set_categories(cats)

    line = LineChart()
    data_line = Reference(ws4, min_col=3, min_row=1, max_row=ws4.max_row)
    line.add_data(data_line, titles_from_data=True)
    line.y_axis.title = "Tỷ lệ (%)"
    line.y_axis.axId = 200
    chart.y_axis.crosses = "max"
    chart += line
    ws4.add_chart(chart, "E5")

    # =============================== #
    # 📊 SHEET 5: GHI CHÚ & CHỮ KÝ
    # =============================== #
    ws5 = wb.create_sheet("Ghi chú & Chữ ký")
    ws5["A1"] = "Ghi chú:"
    ws5["A2"] = "• Báo cáo được xuất tự động từ hệ thống CVD-App."
    ws5["A3"] = "• Dữ liệu cập nhật đến thời điểm xuất file."
    ws5["A5"] = "Người lập báo cáo:"
    ws5["A6"] = session.get('user', 'Quản trị viên')
    ws5["A8"] = "Chữ ký:"
    ws5["A9"] = "____________________________"

    ws5["A1"].font = Font(bold=True, color="1F4E78", size=13)
    ws5.column_dimensions["A"].width = 70

    # =============================== #
    # 💾 XUẤT FILE EXCEL
    # =============================== #
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"ThongKe_CVDApp_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ==========================================
# 🌿 Trang Kiến thức Y học (cho bệnh nhân)
# ==========================================
@app.route('/tips')
def tips():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    # Chỉ cho phép bệnh nhân xem
    if session.get('role') != 'patient':
        flash("Chỉ bệnh nhân mới được truy cập trang này.", "warning")
        return redirect(url_for('home'))
    
    return render_template('tips.html')

# ============================================
# 🤖 API CHAT AI (AJAX) — Nâng cấp chuyên nghiệp
# ============================================
@app.route('/chat_ai_api', methods=['POST'])
def chat_ai_api():
    if 'user' not in session or session.get('role') != 'patient':
        return jsonify({'reply': '⚠️ Bạn chưa đăng nhập hoặc không có quyền truy cập.'}), 403

    import google.generativeai as genai
    from datetime import datetime
    from flask import jsonify

    data = request.get_json()
    msg = data.get('message', '').strip()
    if not msg:
        return jsonify({'reply': 'Vui lòng nhập câu hỏi của bạn.'})

    try:
        # --- Cấu hình model (đã cấu hình sẵn API KEY ở đầu file) ---
        model = genai.GenerativeModel(MODEL_NAME)

        # --- Prompt chuyên nghiệp ---
        prompt = f"""
        Bạn là **Trợ lý y tế ảo CVD-AI**, chuyên tư vấn về **bệnh tim mạch, huyết áp, tiểu đường, lối sống lành mạnh**.
        - Trả lời ngắn gọn, rõ ràng, dễ hiểu, dùng tiếng Việt tự nhiên.
        - Giữ giọng văn **thân thiện, chuyên nghiệp**, tránh dùng từ ngữ phức tạp y học.
        - Nếu câu hỏi ngoài chủ đề sức khỏe, hãy nói nhẹ nhàng: 
          “Xin lỗi, tôi chỉ có thể tư vấn về sức khỏe và tim mạch thôi nhé ”
        - Có thể chia câu trả lời thành 2-3 đoạn rõ ràng.
        - Nếu liên quan đến thói quen, gợi ý **thực hành cụ thể** (ví dụ: “Nên tập thể dục 30 phút mỗi ngày”).
        - Không dùng markdown nặng (chỉ gạch đầu dòng, emoji nhẹ).
        
        Nhiệm vụ: Trả lời câu hỏi của bệnh nhân.
        Câu hỏi từ bệnh nhân: 
        {msg}
        """

        # --- Gọi Gemini API ---
        response = model.generate_content(prompt)

        answer = response.text.strip() if response and response.text else (
            "🤔 Xin lỗi, tôi chưa hiểu rõ câu hỏi của bạn. Bạn có thể diễn đạt lại được không?"
        )

        # --- Làm đẹp phản hồi: xử lý format nhẹ ---
        formatted_answer = (
            answer.replace("**", "")  # bỏ markdown đậm
                    .replace("* ", "• ")  # thay bullet
                    .replace("#", "")
        )

        # --- Lưu vào cơ sở dữ liệu ---
        user_id = session.get('user_id')
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO TinNhanAI (BenhNhanID, NoiDung, PhanHoi, ThoiGian)
            VALUES (?, ?, ?, ?)
        """, (user_id, msg, formatted_answer, datetime.now()))
        conn.commit()
        conn.close()

        # --- Trả về kết quả cho giao diện ---
        return jsonify({'reply': formatted_answer})

    except Exception as e:
        print("⚠️ Lỗi Gemini AI:", e)
        return jsonify({
            'reply': '🚫 Hệ thống AI đang bận hoặc kết nối không ổn định. Vui lòng thử lại sau ít phút.'
        })

# ==========================================
# 📜 API lấy lịch sử chat AI của người dùng hiện tại
# ==========================================
@app.route('/chat_ai_history', methods=['GET'])
def chat_ai_history():
    if 'user' not in session or session.get('role') != 'patient':
        return jsonify({'messages': []}), 403

    user_id = session['user_id']
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT NoiDung, PhanHoi, FORMAT(ThoiGian, 'HH:mm dd/MM') AS ThoiGian
        FROM TinNhanAI
        WHERE BenhNhanID = ?
        ORDER BY ThoiGian
    """, (user_id,))
    rows = cur.fetchall()
    conn.close()

    messages = [
        {'user': r.NoiDung, 'ai': r.PhanHoi, 'time': r.ThoiGian}
        for r in rows
    ]
    return jsonify({'messages': messages})

# ==========================================
# Main
# ==========================================
if __name__ == "__main__":
    app.run(debug=True)
